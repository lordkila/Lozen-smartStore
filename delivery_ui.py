import asyncio
import io
import os

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

os.environ[
    'QT_QPA_PLATFORM_PLUGIN_PATH'] = r'C:\Users\cch23\Desktop\자동화\logen_batch_delivery\vnev\lib\site-packages\PyQt5\Qt5\plugins\platforms'

import sys
from datetime import datetime
from pathlib import Path

import msoffcrypto
import pandas as pd
from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QGroupBox,
    QPushButton, QProgressBar, QTableWidget, QTableWidgetItem, QLabel,
    QFileDialog, QMessageBox, QInputDialog, QPlainTextEdit, QLineEdit
)
from qasync import QEventLoop, asyncSlot


# 엑셀 파일 읽기 함수 (암호 지원)
def read_excel_with_password(file_path, password=None):
    """win32com을 사용하여 암호화된 엑셀 파일 처리"""

    try:
        # 암호가 없는 경우 일반적으로 읽기
        if password is None:
            return pd.read_excel(file_path)

        decrypted_workbook = io.BytesIO()
        office_file = msoffcrypto.OfficeFile(open(file_path, 'rb'))
        office_file.load_key(password=password)
        office_file.decrypt(decrypted_workbook)
        decrypted_workbook.seek(0)
        return pd.read_excel(decrypted_workbook)

    except Exception as e:
        raise ValueError(f"파일 읽기 오류: {str(e)}")


class DragDropWidget(QWidget):
    fileDropped = pyqtSignal(str, str)  # file_path, file_type

    def __init__(self, label_text, file_type):
        super().__init__()
        self.file_type = file_type
        self.file_path = ""
        self.initUI(label_text)

    def initUI(self, label_text):
        layout = QVBoxLayout()

        self.drop_area = QLabel()
        self.drop_area.setText(f'\n\n{label_text}\n파일을 드래그하거나 클릭하여 파일을 선택하세요\n\n')
        self.drop_area.setAlignment(Qt.AlignCenter)
        self.drop_area.setStyleSheet('''
            QLabel {
                border: 2px dashed #4A90E2;
                border-radius: 10px;
                background-color: #F8F9FA;
                color: #666;
                font-size: 14px;
                min-height: 120px;
            }
            QLabel:hover {
                background-color: #E3F2FD;
                cursor: pointer;
            }
        ''')
        self.drop_area.setAcceptDrops(True)
        self.drop_area.dragEnterEvent = self.dragEnterEvent
        self.drop_area.dropEvent = self.dropEvent
        self.drop_area.mousePressEvent = self.mousePressEvent

        self.file_info = QLabel("파일이 선택되지 않았습니다")
        self.file_info.setStyleSheet("color: #666; font-size: 12px;")

        layout.addWidget(self.drop_area)
        layout.addWidget(self.file_info)
        self.setLayout(layout)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if len(urls) == 1:
                file_path = urls[0].toLocalFile()
                if file_path.endswith(('.xlsx', '.xls')):
                    event.accept()
                    self.drop_area.setStyleSheet('''
                        QLabel {
                            border: 2px solid #4A90E2;
                            border-radius: 10px;
                            background-color: #E3F2FD;
                            color: #4A90E2;
                            font-size: 14px;
                            min-height: 120px;
                        }
                    ''')
                    return
        event.ignore()

    def dropEvent(self, event):
        files = [u.toLocalFile() for u in event.mimeData().urls()]
        if files and files[0].endswith(('.xlsx', '.xls')):
            self.handleFileSelected(files[0])
            event.accept()
        else:
            event.ignore()

    def mousePressEvent(self, event):
        """클릭 시 파일 선택 다이얼로그 열기"""
        file_dialog = QFileDialog()
        file_dialog.setFileMode(QFileDialog.ExistingFile)
        file_dialog.setNameFilters(["Excel Files (*.xlsx *.xls)"])

        if file_dialog.exec_():
            selected_files = file_dialog.selectedFiles()
            if selected_files:
                self.handleFileSelected(selected_files[0])

    def handleFileSelected(self, file_path):
        """파일 선택 처리 공통 함수"""
        self.file_path = file_path
        file_name = os.path.basename(self.file_path)
        self.file_info.setText(f"선택된 파일: {file_name}")
        self.drop_area.setText(f'\n\n✓ {file_name}\n파일이 선택되었습니다\n\n')
        self.drop_area.setStyleSheet('''
            QLabel {
                border: 2px solid #4CAF50;
                border-radius: 10px;
                background-color: #E8F5E8;
                color: #4CAF50;
                font-size: 14px;
                min-height: 120px;
            }
        ''')
        self.fileDropped.emit(self.file_path, self.file_type)


class SmartStoreProcessor(QMainWindow):
    def __init__(self):
        super().__init__()
        self.loop = QEventLoop(self)
        asyncio.set_event_loop(self.loop)

        self.a_file_path = ""
        self.b_file_path = ""

        self.initUI()

    def initUI(self):
        self.setWindowTitle('스마트스토어 엑셀 일괄처리 프로그램 v1.2.0')
        self.setGeometry(100, 100, 1000, 700)
        self.setStyleSheet("background-color: #FCFCFC;")

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QVBoxLayout(central_widget)

        # 파일 업로드 섹션
        self.createFileUploadSection(layout)

        # 처리 버튼 섹션
        self.createProcessSection(layout)

        # 결과 섹션
        self.createResultSection(layout)

    def createFileUploadSection(self, parent_layout):
        upload_group = QGroupBox("스마트스토어 엑셀 올리기")
        upload_layout = QHBoxLayout()

        # A 엑셀 (주문 데이터)
        self.a_drop_widget = DragDropWidget("주문 데이터 엑셀 (A)", "order")
        self.a_drop_widget.fileDropped.connect(self.on_file_dropped)

        # B 엑셀 (운송장 데이터)
        self.b_drop_widget = DragDropWidget("운송장 데이터 엑셀 (B)", "shipping")
        self.b_drop_widget.fileDropped.connect(self.on_file_dropped)

        upload_layout.addWidget(self.a_drop_widget)
        upload_layout.addWidget(self.b_drop_widget)

        upload_group.setLayout(upload_layout)
        parent_layout.addWidget(upload_group)

    def createProcessSection(self, parent_layout):
        process_group = QGroupBox("처리 옵션")
        process_layout = QHBoxLayout()

        self.process_button = QPushButton("일괄처리 시작")
        self.process_button.setStyleSheet('''
            QPushButton {
                background-color: #4A90E2;
                color: white;
                border-radius: 10px;
                padding: 12px 24px;
                font-size: 16px;
                font-weight: bold;
                min-width: 200px;
            }
            QPushButton:hover {
                background-color: #357ABD;
            }
            QPushButton:disabled {
                background-color: #CCCCCC;
                color: #666666;
            }
        ''')
        self.process_button.clicked.connect(self.start_processing)
        self.process_button.setEnabled(False)

        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)

        process_layout.addWidget(self.process_button)
        process_layout.addWidget(self.progress_bar)
        process_layout.addStretch()

        process_group.setLayout(process_layout)
        parent_layout.addWidget(process_group)

    def createResultSection(self, parent_layout):
        result_group = QGroupBox("처리 결과 및 로그")
        result_layout = QVBoxLayout()

        # 결과 테이블
        self.result_table = QTableWidget(0, 9)
        self.result_table.setHorizontalHeaderLabels([
            "상품주문번호", "배송방법", "택배사", "송장번호",
            "상품명", "수량", "수취인", "수취인연락처", "배송지"
        ])
        self.result_table.horizontalHeader().setStretchLastSection(True)

        # 로그 영역
        self.log_area = QPlainTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(200)
        self.log_area.setStyleSheet('''
            QPlainTextEdit {
                background-color: #454545;
                color: white;
                font-family: "Consolas", "Monaco", monospace;
                font-size: 12px;
            }
        ''')

        result_layout.addWidget(self.result_table)
        result_layout.addWidget(self.log_area)

        result_group.setLayout(result_layout)
        parent_layout.addWidget(result_group)

    def on_file_dropped(self, file_path, file_type):
        if file_type == "order":
            self.a_file_path = file_path
            self.log(f"주문 데이터 파일 선택됨: {os.path.basename(file_path)}")
        elif file_type == "shipping":
            self.b_file_path = file_path
            self.log(f"운송장 데이터 파일 선택됨: {os.path.basename(file_path)}")

        # 두 파일이 모두 선택되면 처리 버튼 활성화
        if self.a_file_path and self.b_file_path:
            self.process_button.setEnabled(True)
            self.log("두 파일이 모두 선택되었습니다. 처리를 시작할 수 있습니다.")

    def log(self, message):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_area.appendPlainText(f"[{timestamp}] {message}")

    def start_processing(self):
        if not self.a_file_path or not self.b_file_path:
            QMessageBox.warning(self, "경고", "두 파일을 모두 선택해주세요.")
            return

        self.process_button.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)  # 무한 프로그레스

        # 비동기 처리 시작
        asyncio.ensure_future(self.process_files())

    def get_password(self, file_name):
        """암호 입력 다이얼로그 표시 (최대 3회 시도)"""
        attempts = 0
        while attempts < 3:
            password, ok = QInputDialog.getText(
                self,
                "암호 입력",
                f"파일에 암호가 걸려 있습니다. \n\n암호를 입력하세요\t (시도 {attempts + 1}/3):",
                QLineEdit.Password
            )
            if not ok:
                return None
            return password

        QMessageBox.critical(self, "오류", "3회 이상 암호 입력에 실패했습니다.")
        return None

    @asyncSlot()
    async def process_files(self):
        try:
            self.log("파일 처리를 시작합니다...")

            # A 엑셀 읽기 (주문 데이터)
            self.log(f"주문 데이터 읽는 중: {os.path.basename(self.a_file_path)}")

            # 암호 처리 로직
            password = None
            try:
                # 먼저 암호 없이 시도
                a_df = read_excel_with_password(self.a_file_path)
            except Exception as e:
                # 암호가 필요한 경우
                if "ole2" in str(e).lower() or "password" in str(e).lower() or "encrypted" in str(e).lower():
                    password = self.get_password(os.path.basename(self.a_file_path))
                    if password is None:
                        self.log("암호 입력 취소로 처리를 중단합니다.")
                        return
                    a_df = read_excel_with_password(self.a_file_path, password)
                else:
                    raise e
            a_df.columns = a_df.iloc[0]
            a_df.iloc[1:].reset_index(drop=True)
            self.log(f"주문 데이터 {len(a_df)}행 로드됨")

            # B 엑셀 읽기 (운송장 데이터)
            self.log(f"운송장 데이터 읽는 중: {os.path.basename(self.b_file_path)}")

            b_df = pd.read_excel(self.b_file_path)
            b_df.columns = b_df.iloc[1]
            b_df.iloc[2:].reset_index(drop=True)
            self.log(f"운송장 데이터 {len(b_df)}행 로드됨")

            # 데이터 매칭 처리
            self.log("데이터 매칭 중...")
            result_rows = []

            # fixme test
            # r_a = []
            # r_b = []
            # r_c = []

            for _, a_row in a_df.iterrows():
                try:
                    a_name = str(a_row['수취인명']).strip()
                    a_phone = str(a_row['수취인연락처1']).strip()
                    a_addr = str(a_row['통합배송지']).strip()

                    for _, b_row in b_df.iterrows():
                        try:
                            b_name = str(b_row['수하인명']).strip()
                            b_phone = str(b_row['수하인전화']).replace('*', '').strip()
                            b_addr = str(b_row['수하인주소1']).strip()
                            b_addr_words = b_addr.split(' ')

                            # fixme test
                            # if b_name == a_name:
                            #     r_a.append(a_name)
                            # if b_phone in a_phone:
                            #     r_b.append(a_name)
                            # if len(b_addr_words) > 2 and b_addr_words[1] in a_addr:
                            #     r_c.append(a_name)

                            # 매칭 조건 확인
                            if (b_name == a_name and
                                    b_phone in a_phone and
                                    len(b_addr_words) > 2 and
                                    b_addr_words[1] in a_addr):
                                result_rows.append({
                                    '상품주문번호': a_row['상품주문번호'],
                                    '배송방법': a_row['배송방법'],
                                    '택배사': a_row['택배사'],
                                    '송장번호': b_row['운송장번호'],
                                    '상품명': a_row['상품명'],
                                    '수량': a_row['수량'],
                                    '수취인': a_row['수취인명'],
                                    '수취인연락처': a_row['수취인연락처1'],
                                    '배송지': a_row['통합배송지']
                                })
                                break
                        except Exception as e:
                            continue
                except Exception as e:
                    continue

            self.log(f"매칭 완료: {len(result_rows)}건 처리됨")

            # fixme test
            # self.log(f"{len(r_a)}, {len(r_b)}, {len(r_c)}")

            # 결과를 테이블에 표시
            self.display_results(result_rows)

            # 엑셀 파일로 저장
            if result_rows:
                await self.save_to_excel(result_rows)
            else:
                self.log("매칭된 데이터가 없습니다.")

        except Exception as e:
            self.log(f"처리 중 오류 발생: {str(e)}")
            QMessageBox.critical(self, "오류", f"처리 중 오류가 발생했습니다:\n{str(e)}")
        finally:
            self.progress_bar.setVisible(False)
            self.process_button.setEnabled(True)

    def display_results(self, result_rows):
        self.result_table.setRowCount(len(result_rows))

        for row_idx, row_data in enumerate(result_rows):
            for col_idx, (key, value) in enumerate(row_data.items()):
                item = QTableWidgetItem(str(value))
                self.result_table.setItem(row_idx, col_idx, item)

    async def save_to_excel(self, result_rows):
        try:
            a_file_name = Path(self.a_file_path).stem
            output_filename = f"일괄처리_{a_file_name}.xlsx"
            df = pd.DataFrame(result_rows)
            sheet_name = '발송처리'

            with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            wb = load_workbook(output_filename)
            ws = wb[sheet_name]

            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            for cell in ws[1]:  # 첫 번째 행 (헤더)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 열 너비 자동 조정
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 5, 200)
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(output_filename)

            self.log(f"결과 파일 저장 완료: {output_filename}")

            # 저장 완료 메시지
            reply = QMessageBox.information(
                self, "완료",
                f"처리가 완료되었습니다!\n\n"
                f"저장된 파일: {output_filename}\n"
                f"처리된 건수: {len(result_rows)}건\n\n"
                f"파일을 열어보시겠습니까?",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                os.startfile(output_filename)

        except Exception as e:
            self.log(f"파일 저장 중 오류: {str(e)}")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    loop = QEventLoop(app)

    # Qt 플러그인 경로 설정 (Windows에서 필요한 경우)
    if sys.platform == 'win32':
        import PyQt5

        plugin_path = os.path.join(
            os.path.dirname(PyQt5.__file__),
            'Qt', 'plugins', 'platforms'
        )
        os.environ['QT_QPA_PLATFORM_PLUGIN_PATH'] = plugin_path

    window = SmartStoreProcessor()
    window.show()

    with loop:
        loop.run_forever()
